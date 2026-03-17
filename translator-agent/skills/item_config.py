# skills/item_config.py
"""
Item配置管理技能 - 管理4个独立的Excel文件（无备份）
"""

import openpyxl
import os

class ItemConfigSkill:
    """Item配置管理技能 - 管理4个独立的Excel文件"""
    
    def __init__(self, config_dir="config"):
        """
        初始化技能
        
        Args:
            config_dir: 配置文件目录
        """
        self.skill_description = {
            "name": "item_config",
            "description": "管理游戏Item配置（4个独立Excel文件）",
            "version": "1.0.0"
        }
        
        # 配置文件目录
        self.config_dir = os.path.abspath(config_dir)
        os.makedirs(self.config_dir, exist_ok=True)
        
        # 4个独立文件的路径
        self.files = {
            "Items": os.path.join(self.config_dir, "Items.xlsx"),
            "PriceConfig": os.path.join(self.config_dir, "PriceConfig.xlsx"),
            "DropConfig": os.path.join(self.config_dir, "DropConfig.xlsx"),
            "RecipeConfig": os.path.join(self.config_dir, "RecipeConfig.xlsx")
        }
        
        print(f"📁 配置文件目录: {self.config_dir}")
        for name, path in self.files.items():
            exists = "✅" if os.path.exists(path) else "❌"
            print(f"  {exists} {name}: {os.path.basename(path)}")
        
        # 检查并创建缺失的文件
        self._ensure_all_files_exist()
    
    def _ensure_all_files_exist(self):
        """确保所有4个文件都存在，不存在则创建"""
        for sheet_name, file_path in self.files.items():
            if not os.path.exists(file_path):
                self._create_empty_file(sheet_name, file_path)
    
    def _create_empty_file(self, sheet_name, file_path):
        """创建空的Excel文件"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 根据表名设置不同的表头
        headers = {
            "Items": ["ID", "名称", "类型", "攻击力", "描述"],
            "PriceConfig": ["ItemID", "名称", "购买价", "出售价"],
            "DropConfig": ["ItemID", "名称", "掉落概率", "掉落数量"],
            "RecipeConfig": ["ItemID", "名称", "材料1", "材料2"]
        }
        
        # 写入表头
        for col, header in enumerate(headers[sheet_name], 1):
            ws.cell(row=1, column=col, value=header)
        
        wb.save(file_path)
        print(f"✅ 已创建文件: {os.path.basename(file_path)}")
    
    def _get_next_item_id(self):
        """从Items文件获取下一个可用的Item ID"""
        file_path = self.files["Items"]
        
        if not os.path.exists(file_path):
            return "item_001"
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        max_num = 0
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and str(cell_value).startswith("item_"):
                try:
                    num = int(str(cell_value).split("_")[1])
                    max_num = max(max_num, num)
                except:
                    continue
        
        return f"item_{str(max_num + 1).zfill(3)}"
    
    def _get_type_defaults(self, item_type):
        """根据物品类型获取默认值"""
        defaults = {
            "武器": {
                "购买价": 100,
                "出售价": 50,
                "掉落概率": 0.05,
                "掉落数量": 1,
                "攻击力": 10
            },
            "防具": {
                "购买价": 80,
                "出售价": 40,
                "掉落概率": 0.08,
                "掉落数量": 1,
                "防御力": 10
            },
            "消耗品": {
                "购买价": 20,
                "出售价": 10,
                "掉落概率": 0.3,
                "掉落数量": 2,
                "恢复量": 50
            },
            "材料": {
                "购买价": 5,
                "出售价": 2,
                "掉落概率": 0.5,
                "掉落数量": 3,
            },
            "任务物品": {
                "购买价": 0,
                "出售价": 0,
                "掉落概率": 0.2,
                "掉落数量": 1,
            }
        }
        return defaults.get(item_type, {
            "购买价": 50,
            "出售价": 25,
            "掉落概率": 0.1,
            "掉落数量": 1
        })
    
    def add_item(self, name, item_type, **kwargs):
        """
        新增Item，自动更新所有4个文件
        """
        param_mapping = {
        'attack': '攻击力',
        'description': '描述',
        'drop_rate': '掉落概率',
        'drop_count': '掉落数量',
        'buy_price': '购买价',
        'sell_price': '出售价',
        'material1': '材料1',
        'material2': '材料2'
    }
        for eng_key, chn_key in param_mapping.items():
            if eng_key in kwargs and chn_key not in kwargs:
                kwargs[chn_key] = kwargs[eng_key]
        print("\n" + "=" * 60)
        print("🔍 add_item 收到的参数:")
        print(f"  name: {name} (类型: {type(name)})")
        print(f"  item_type: {item_type} (类型: {type(item_type)})")
        print(f"  kwargs: {kwargs}")
        
        # 分别打印各个关键参数
        print("\n📦 关键参数值:")
        print(f"  attack: {kwargs.get('attack')} / {kwargs.get('攻击力')}")
        print(f"  description: {kwargs.get('description')} / {kwargs.get('描述')}")
        print(f"  drop_rate: {kwargs.get('drop_rate')} / {kwargs.get('掉落概率')}")
        print(f"  drop_count: {kwargs.get('drop_count')} / {kwargs.get('掉落数量')}")
        print(f"  buy_price: {kwargs.get('buy_price')} / {kwargs.get('购买价')}")
        print(f"  sell_price: {kwargs.get('sell_price')} / {kwargs.get('出售价')}")
        print(f"  material1: {kwargs.get('material1')} / {kwargs.get('材料1')}")
        print(f"  material2: {kwargs.get('material2')} / {kwargs.get('材料2')}")
        print("=" * 60 + "\n")
        
        try:
            print(f"🔧 开始添加物品: {name}, 类型: {item_type}")
            
            # 1. 生成新ID
            new_id = self._get_next_item_id()
            
            # 2. 获取类型默认值
            defaults = self._get_type_defaults(item_type)
            
            results = [f"📦 正在添加新物品: {name} ({item_type})"]
            results.append(f"🆔 生成ID: {new_id}")
            
            # ===== 更新Items文件 =====
            items_file = self.files["Items"]
            print(f"  写入 Items 文件: {items_file}")
            
            wb_items = openpyxl.load_workbook(items_file)
            ws_items = wb_items.active
            
            next_row = ws_items.max_row + 1
            ws_items.cell(row=next_row, column=1, value=new_id)
            ws_items.cell(row=next_row, column=2, value=name)
            ws_items.cell(row=next_row, column=3, value=item_type)
            
            attack = kwargs.get("attack", kwargs.get("攻击力", defaults.get("攻击力", 0)))
            ws_items.cell(row=next_row, column=4, value=attack)
            
            description = kwargs.get("description", kwargs.get("描述", f"一{'' if item_type=='武器' else '件'}{item_type}"))
            ws_items.cell(row=next_row, column=5, value=description)
            
            # 保存
            wb_items.save(items_file)
            print(f"  ✅ Items文件已保存")
            
            # 立即验证
            wb_items_check = openpyxl.load_workbook(items_file)
            ws_check = wb_items_check.active
            last_row_value = ws_check.cell(row=next_row, column=2).value
            print(f"  ✅ 验证: Items文件最后一行名称 = {last_row_value}")
            
            results.append(f"✅ Items文件已更新: {os.path.basename(items_file)}")
            
            # ===== 更新PriceConfig文件 =====
            price_file = self.files["PriceConfig"]
            print(f"  写入 PriceConfig 文件: {price_file}")
            
            wb_price = openpyxl.load_workbook(price_file)
            ws_price = wb_price.active
            
            next_row = ws_price.max_row + 1
            ws_price.cell(row=next_row, column=1, value=new_id)
            ws_price.cell(row=next_row, column=2, value=name)
            
            buy_price = kwargs.get("buy_price", kwargs.get("购买价", defaults["购买价"]))
            ws_price.cell(row=next_row, column=3, value=buy_price)
            
            sell_price = kwargs.get("sell_price", kwargs.get("出售价", buy_price // 2))
            ws_price.cell(row=next_row, column=4, value=sell_price)
            
            wb_price.save(price_file)
            print(f"  ✅ PriceConfig文件已保存")
            
            # 验证
            wb_price_check = openpyxl.load_workbook(price_file)
            ws_price_check = wb_price_check.active
            last_row_value = ws_price_check.cell(row=next_row, column=2).value
            print(f"  ✅ 验证: PriceConfig文件最后一行名称 = {last_row_value}")
            
            results.append(f"✅ PriceConfig文件已更新: {os.path.basename(price_file)}")
            
            # ===== 更新DropConfig文件 =====
            drop_file = self.files["DropConfig"]
            print(f"  写入 DropConfig 文件: {drop_file}")
            
            wb_drop = openpyxl.load_workbook(drop_file)
            ws_drop = wb_drop.active
            
            next_row = ws_drop.max_row + 1
            ws_drop.cell(row=next_row, column=1, value=new_id)
            ws_drop.cell(row=next_row, column=2, value=name)
            
            drop_rate = kwargs.get("drop_rate", kwargs.get("掉落概率", defaults["掉落概率"]))
            ws_drop.cell(row=next_row, column=3, value=drop_rate)
            
            drop_count = kwargs.get("drop_count", kwargs.get("掉落数量", defaults["掉落数量"]))
            ws_drop.cell(row=next_row, column=4, value=drop_count)
            
            wb_drop.save(drop_file)
            print(f"  ✅ DropConfig文件已保存")
            
            # 验证
            wb_drop_check = openpyxl.load_workbook(drop_file)
            ws_drop_check = wb_drop_check.active
            last_row_value = ws_drop_check.cell(row=next_row, column=2).value
            print(f"  ✅ 验证: DropConfig文件最后一行名称 = {last_row_value}")
            
            results.append(f"✅ DropConfig文件已更新: {os.path.basename(drop_file)}")
            
            # ===== 更新RecipeConfig文件 =====
            recipe_file = self.files["RecipeConfig"]
            print(f"  写入 RecipeConfig 文件: {recipe_file}")
            
            wb_recipe = openpyxl.load_workbook(recipe_file)
            ws_recipe = wb_recipe.active
            
            next_row = ws_recipe.max_row + 1
            ws_recipe.cell(row=next_row, column=1, value=new_id)
            ws_recipe.cell(row=next_row, column=2, value=name)
            
            material1 = kwargs.get("material1", kwargs.get("材料1", ""))
            material2 = kwargs.get("material2", kwargs.get("材料2", ""))
            ws_recipe.cell(row=next_row, column=3, value=material1)
            ws_recipe.cell(row=next_row, column=4, value=material2)
            
            wb_recipe.save(recipe_file)
            print(f"  ✅ RecipeConfig文件已保存")
            
            # 验证
            wb_recipe_check = openpyxl.load_workbook(recipe_file)
            ws_recipe_check = wb_recipe_check.active
            last_row_value = ws_recipe_check.cell(row=next_row, column=2).value
            print(f"  ✅ 验证: RecipeConfig文件最后一行名称 = {last_row_value}")
            
            results.append(f"✅ RecipeConfig文件已更新: {os.path.basename(recipe_file)}")
            
            results.append(f"🎉 所有4个文件更新完成！")
            
            return "\n".join(results)
            
        except Exception as e:
            import traceback
            error_msg = f"❌ 操作失败: {str(e)}\n"
            error_msg += traceback.format_exc()
            print(error_msg)
            return f"❌ 操作失败: {str(e)}"
    
    def query_item(self, item_id_or_name):
        """
        查询Item的完整配置（从4个文件中）
        
        Args:
            item_id_or_name: Item ID 或 名称
        
        Returns:
            str: 查询结果
        """
        try:
            results = [f"🔍 查询: {item_id_or_name}"]
            
            # 从Items文件查找ID
            item_id = None
            item_name = None
            
            items_file = self.files["Items"]
            if os.path.exists(items_file):
                wb_items = openpyxl.load_workbook(items_file)
                ws_items = wb_items.active
                
                for row in range(2, ws_items.max_row + 1):
                    id_val = ws_items.cell(row=row, column=1).value
                    name_val = ws_items.cell(row=row, column=2).value
                    
                    if id_val == item_id_or_name or name_val == item_id_or_name:
                        item_id = id_val
                        item_name = name_val
                        
                        # 读取Items表数据
                        item_data = []
                        for col in range(1, ws_items.max_column + 1):
                            val = ws_items.cell(row=row, column=col).value
                            item_data.append(str(val) if val else "-")
                        
                        results.append(f"\n📄 **Items.xlsx**:")
                        results.append("  " + " | ".join(item_data))
                        break
            
            if not item_id:
                return f"❌ 未找到Item: {item_id_or_name}"
            
            # 从PriceConfig查询
            price_file = self.files["PriceConfig"]
            if os.path.exists(price_file):
                wb_price = openpyxl.load_workbook(price_file)
                ws_price = wb_price.active
                
                for row in range(2, ws_price.max_row + 1):
                    if ws_price.cell(row=row, column=1).value == item_id:
                        price_data = []
                        for col in range(1, ws_price.max_column + 1):
                            val = ws_price.cell(row=row, column=col).value
                            price_data.append(str(val) if val else "-")
                        
                        results.append(f"\n📄 **PriceConfig.xlsx**:")
                        results.append("  " + " | ".join(price_data))
                        break
            
            # 从DropConfig查询
            drop_file = self.files["DropConfig"]
            if os.path.exists(drop_file):
                wb_drop = openpyxl.load_workbook(drop_file)
                ws_drop = wb_drop.active
                
                for row in range(2, ws_drop.max_row + 1):
                    if ws_drop.cell(row=row, column=1).value == item_id:
                        drop_data = []
                        for col in range(1, ws_drop.max_column + 1):
                            val = ws_drop.cell(row=row, column=col).value
                            drop_data.append(str(val) if val else "-")
                        
                        results.append(f"\n📄 **DropConfig.xlsx**:")
                        results.append("  " + " | ".join(drop_data))
                        break
            
            # 从RecipeConfig查询
            recipe_file = self.files["RecipeConfig"]
            if os.path.exists(recipe_file):
                wb_recipe = openpyxl.load_workbook(recipe_file)
                ws_recipe = wb_recipe.active
                
                for row in range(2, ws_recipe.max_row + 1):
                    if ws_recipe.cell(row=row, column=1).value == item_id:
                        recipe_data = []
                        for col in range(1, ws_recipe.max_column + 1):
                            val = ws_recipe.cell(row=row, column=col).value
                            recipe_data.append(str(val) if val else "-")
                        
                        results.append(f"\n📄 **RecipeConfig.xlsx**:")
                        results.append("  " + " | ".join(recipe_data))
                        break
            
            return "\n".join(results)
            
        except Exception as e:
            return f"❌ 查询失败: {str(e)}"
    
    def list_items(self, item_type=None):
        """
        列出所有Item（从Items文件）
        
        Args:
            item_type: 筛选类型
        
        Returns:
            str: Item列表
        """
        try:
            items_file = self.files["Items"]
            if not os.path.exists(items_file):
                return "📭 Items文件不存在"
            
            wb_items = openpyxl.load_workbook(items_file)
            ws_items = wb_items.active
            
            items = []
            for row in range(2, ws_items.max_row + 1):
                item_id = ws_items.cell(row=row, column=1).value
                name = ws_items.cell(row=row, column=2).value
                type_ = ws_items.cell(row=row, column=3).value
                attack = ws_items.cell(row=row, column=4).value
                
                if not item_id:  # 跳过空行
                    continue
                    
                if item_type and type_ != item_type:
                    continue
                
                items.append(f"  {item_id} | {name} | {type_} | 攻击力:{attack}")
            
            if not items:
                return f"📭 未找到任何Item" + (f" (类型: {item_type})" if item_type else "")
            
            header = f"📋 Item列表 ({os.path.basename(items_file)})" + (f" - 类型: {item_type}" if item_type else "")
            return header + "\n" + "\n".join(items)
            
        except Exception as e:
            return f"❌ 列出Item失败: {str(e)}"
    
    def update_item(self, item_id_or_name, **updates):
        """
        更新Item的特定属性
        """
        try:
            # 先找到Item ID
            item_id = None
            items_file = self.files["Items"]
            wb_items = openpyxl.load_workbook(items_file)
            ws_items = wb_items.active
            
            for row in range(2, ws_items.max_row + 1):
                if ws_items.cell(row=row, column=1).value == item_id_or_name:
                    item_id = item_id_or_name
                    break
                if ws_items.cell(row=row, column=2).value == item_id_or_name:
                    item_id = ws_items.cell(row=row, column=1).value
                    break
            
            if not item_id:
                return f"❌ 未找到Item: {item_id_or_name}"
            
            updated = []
            
            # 更新Items文件
            if any(k in updates for k in ["名称", "类型", "攻击力", "描述"]):
                for row in range(2, ws_items.max_row + 1):
                    if ws_items.cell(row=row, column=1).value == item_id:
                        if "名称" in updates:
                            ws_items.cell(row=row, column=2, value=updates["名称"])
                            updated.append(f"名称={updates['名称']}")
                        if "类型" in updates:
                            ws_items.cell(row=row, column=3, value=updates["类型"])
                            updated.append(f"类型={updates['类型']}")
                        if "攻击力" in updates:
                            ws_items.cell(row=row, column=4, value=updates["攻击力"])
                            updated.append(f"攻击力={updates['攻击力']}")
                        if "描述" in updates:
                            ws_items.cell(row=row, column=5, value=updates["描述"])
                            updated.append(f"描述={updates['描述']}")
                        break
                wb_items.save(items_file)
            
            # 更新PriceConfig文件
            if any(k in updates for k in ["购买价", "出售价"]):
                price_file = self.files["PriceConfig"]
                wb_price = openpyxl.load_workbook(price_file)
                ws_price = wb_price.active
                
                for row in range(2, ws_price.max_row + 1):
                    if ws_price.cell(row=row, column=1).value == item_id:
                        if "购买价" in updates:
                            ws_price.cell(row=row, column=3, value=updates["购买价"])
                            updated.append(f"购买价={updates['购买价']}")
                        if "出售价" in updates:
                            ws_price.cell(row=row, column=4, value=updates["出售价"])
                            updated.append(f"出售价={updates['出售价']}")
                        break
                wb_price.save(price_file)
            
            # 更新DropConfig文件
            if any(k in updates for k in ["掉落概率", "掉落数量"]):
                drop_file = self.files["DropConfig"]
                wb_drop = openpyxl.load_workbook(drop_file)
                ws_drop = wb_drop.active
                
                for row in range(2, ws_drop.max_row + 1):
                    if ws_drop.cell(row=row, column=1).value == item_id:
                        if "掉落概率" in updates:
                            ws_drop.cell(row=row, column=3, value=updates["掉落概率"])
                            updated.append(f"掉落概率={updates['掉落概率']}")
                        if "掉落数量" in updates:
                            ws_drop.cell(row=row, column=4, value=updates["掉落数量"])
                            updated.append(f"掉落数量={updates['掉落数量']}")
                        break
                wb_drop.save(drop_file)
            
            # 更新RecipeConfig文件
            if any(k in updates for k in ["材料1", "材料2"]):
                recipe_file = self.files["RecipeConfig"]
                wb_recipe = openpyxl.load_workbook(recipe_file)
                ws_recipe = wb_recipe.active
                
                for row in range(2, ws_recipe.max_row + 1):
                    if ws_recipe.cell(row=row, column=1).value == item_id:
                        if "材料1" in updates:
                            ws_recipe.cell(row=row, column=3, value=updates["材料1"])
                            updated.append(f"材料1={updates['材料1']}")
                        if "材料2" in updates:
                            ws_recipe.cell(row=row, column=4, value=updates["材料2"])
                            updated.append(f"材料2={updates['材料2']}")
                        break
                wb_recipe.save(recipe_file)
            
            if updated:
                return f"✅ 已更新 {item_id}:\n" + "\n".join(updated)
            else:
                return "没有字段被更新"
            
        except Exception as e:
            return f"❌ 更新失败: {str(e)}"
    
    def delete_item(self, item_id_or_name, confirm=False):
        """
        删除Item（从所有4个文件中）
        """
        if not confirm:
            return f"⚠️ 确认要删除 {item_id_or_name} 吗？这将从所有4个文件中删除！如需删除，请加上 confirm=True"
        
        try:
            # 先找到Item ID
            item_id = None
            items_file = self.files["Items"]
            wb_items = openpyxl.load_workbook(items_file)
            ws_items = wb_items.active
            
            for row in range(2, ws_items.max_row + 1):
                if ws_items.cell(row=row, column=1).value == item_id_or_name:
                    item_id = item_id_or_name
                    break
                if ws_items.cell(row=row, column=2).value == item_id_or_name:
                    item_id = ws_items.cell(row=row, column=1).value
                    break
            
            if not item_id:
                return f"❌ 未找到Item: {item_id_or_name}"
            
            # 从Items文件删除
            for row in range(2, ws_items.max_row + 1):
                if ws_items.cell(row=row, column=1).value == item_id:
                    ws_items.delete_rows(row)
                    break
            wb_items.save(items_file)
            
            # 从PriceConfig文件删除
            price_file = self.files["PriceConfig"]
            wb_price = openpyxl.load_workbook(price_file)
            ws_price = wb_price.active
            
            for row in range(2, ws_price.max_row + 1):
                if ws_price.cell(row=row, column=1).value == item_id:
                    ws_price.delete_rows(row)
                    break
            wb_price.save(price_file)
            
            # 从DropConfig文件删除
            drop_file = self.files["DropConfig"]
            wb_drop = openpyxl.load_workbook(drop_file)
            ws_drop = wb_drop.active
            
            for row in range(2, ws_drop.max_row + 1):
                if ws_drop.cell(row=row, column=1).value == item_id:
                    ws_drop.delete_rows(row)
                    break
            wb_drop.save(drop_file)
            
            # 从RecipeConfig文件删除
            recipe_file = self.files["RecipeConfig"]
            wb_recipe = openpyxl.load_workbook(recipe_file)
            ws_recipe = wb_recipe.active
            
            for row in range(2, ws_recipe.max_row + 1):
                if ws_recipe.cell(row=row, column=1).value == item_id:
                    ws_recipe.delete_rows(row)
                    break
            wb_recipe.save(recipe_file)
            
            return f"✅ 已从所有4个文件中删除: {item_id}"
            
        except Exception as e:
            return f"❌ 删除失败: {str(e)}"
    
    def info(self):
        return self.skill_description