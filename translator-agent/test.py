# check_excel.py
import openpyxl
import os

def check_excel_files():
    """直接查看Excel文件内容"""
    
    config_dir = "config"
    files = ["Items.xlsx", "PriceConfig.xlsx", "DropConfig.xlsx", "RecipeConfig.xlsx"]
    
    print("=" * 60)
    print("🔍 直接查看Excel文件内容")
    print("=" * 60)
    
    for file_name in files:
        file_path = os.path.join(config_dir, file_name)
        print(f"\n📄 {file_name}:")
        
        if not os.path.exists(file_path):
            print(f"   ❌ 文件不存在")
            continue
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # 打印表头
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                headers.append(str(header) if header else "空")
            print(f"   表头: {' | '.join(headers)}")
            
            # 打印数据行
            if ws.max_row > 1:
                for row in range(2, ws.max_row + 1):
                    row_data = []
                    for col in range(1, ws.max_column + 1):
                        val = ws.cell(row=row, column=col).value
                        row_data.append(str(val) if val is not None else "空")
                    print(f"   行{row}: {' | '.join(row_data)}")
            else:
                print("   没有数据")
                
        except Exception as e:
            print(f"   ❌ 读取失败: {e}")

def check_with_query(skill):
    """通过查询接口查看"""
    print("\n" + "=" * 60)
    print("🔍 通过查询接口查看")
    print("=" * 60)
    
    # 列出所有物品
    print("\n📋 所有物品列表:")
    list_result = skill.list_items()
    print(list_result)
    
    # 直接查询屠龙刀
    print("\n🔍 查询屠龙刀:")
    query_result = skill.query_item("屠龙刀")
    print(query_result)

# check_last_rows.py
import openpyxl
import os

def check_last_rows():
    """查看每个文件的最后几行"""
    
    config_dir = "config"
    files = ["Items.xlsx", "PriceConfig.xlsx", "DropConfig.xlsx", "RecipeConfig.xlsx"]
    
    print("=" * 60)
    print("🔍 查看每个文件的最后5行")
    print("=" * 60)
    
    for file_name in files:
        file_path = os.path.join(config_dir, file_name)
        print(f"\n📄 {file_name} 最后5行:")
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        start_row = max(2, ws.max_row - 4)  # 最后5行（从第2行开始）
        for row in range(start_row, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                row_data.append(str(val) if val is not None else "空")
            print(f"  行{row}: {' | '.join(row_data)}")

if __name__ == "__main__":
    check_last_rows()