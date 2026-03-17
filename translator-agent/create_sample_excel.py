# create_sample_files.py
"""
创建4个示例Excel文件
"""

import openpyxl
import os

def create_sample_files():
    """创建4个示例Excel文件"""
    
    config_dir = "config"
    os.makedirs(config_dir, exist_ok=True)
    
    # 1. 创建Items.xlsx
    wb_items = openpyxl.Workbook()
    ws_items = wb_items.active
    ws_items.title = "Items"
    
    headers = ["ID", "名称", "类型", "攻击力", "描述"]
    for col, header in enumerate(headers, 1):
        ws_items.cell(row=1, column=col, value=header)
    
    sample_items = [
        ("item_001", "木剑", "武器", 10, "新手装备"),
        ("item_002", "布衣", "防具", 0, "新手防具"),
        ("item_003", "小红药", "消耗品", 0, "回复50HP"),
    ]
    
    for row, item in enumerate(sample_items, 2):
        for col, value in enumerate(item, 1):
            ws_items.cell(row=row, column=col, value=value)
    
    wb_items.save(os.path.join(config_dir, "Items.xlsx"))
    print("✅ 已创建: Items.xlsx")
    
    # 2. 创建PriceConfig.xlsx
    wb_price = openpyxl.Workbook()
    ws_price = wb_price.active
    ws_price.title = "PriceConfig"
    
    headers = ["ItemID", "名称", "购买价", "出售价"]
    for col, header in enumerate(headers, 1):
        ws_price.cell(row=1, column=col, value=header)
    
    sample_prices = [
        ("item_001", "木剑", 100, 50),
        ("item_002", "布衣", 80, 40),
        ("item_003", "小红药", 20, 10),
    ]
    
    for row, price in enumerate(sample_prices, 2):
        for col, value in enumerate(price, 1):
            ws_price.cell(row=row, column=col, value=value)
    
    wb_price.save(os.path.join(config_dir, "PriceConfig.xlsx"))
    print("✅ 已创建: PriceConfig.xlsx")
    
    # 3. 创建DropConfig.xlsx
    wb_drop = openpyxl.Workbook()
    ws_drop = wb_drop.active
    ws_drop.title = "DropConfig"
    
    headers = ["ItemID", "名称", "掉落概率", "掉落数量"]
    for col, header in enumerate(headers, 1):
        ws_drop.cell(row=1, column=col, value=header)
    
    sample_drops = [
        ("item_001", "木剑", 0.1, 1),
        ("item_002", "布衣", 0.15, 1),
        ("item_003", "小红药", 0.3, 2),
    ]
    
    for row, drop in enumerate(sample_drops, 2):
        for col, value in enumerate(drop, 1):
            ws_drop.cell(row=row, column=col, value=value)
    
    wb_drop.save(os.path.join(config_dir, "DropConfig.xlsx"))
    print("✅ 已创建: DropConfig.xlsx")
    
    # 4. 创建RecipeConfig.xlsx
    wb_recipe = openpyxl.Workbook()
    ws_recipe = wb_recipe.active
    ws_recipe.title = "RecipeConfig"
    
    headers = ["ItemID", "名称", "材料1", "材料2"]
    for col, header in enumerate(headers, 1):
        ws_recipe.cell(row=1, column=col, value=header)
    
    sample_recipes = [
        ("item_001", "木剑", "木材x2", "铁锭x1"),
        ("item_002", "布衣", "布料x3", "线x2"),
    ]
    
    for row, recipe in enumerate(sample_recipes, 2):
        for col, value in enumerate(recipe, 1):
            ws_recipe.cell(row=row, column=col, value=value)
    
    wb_recipe.save(os.path.join(config_dir, "RecipeConfig.xlsx"))
    print("✅ 已创建: RecipeConfig.xlsx")
    
    print("\n🎉 所有4个文件创建完成！")
    print(f"📁 位置: {os.path.abspath(config_dir)}")

if __name__ == "__main__":
    create_sample_files()