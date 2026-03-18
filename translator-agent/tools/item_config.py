# tools/item_tools.py
import openpyxl
import os
from pathlib import Path

class ItemTools:
    """物品管理工具"""
    
    def __init__(self, config_dir="config"):
        self.config_dir = config_dir
        os.makedirs(config_dir, exist_ok=True)
        self.items_file = os.path.join(config_dir, "Items.xlsx")
        self._init_file()
    
    def _init_file(self):
        """初始化Excel文件"""
        if not os.path.exists(self.items_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Items"
            headers = ["ID", "名称", "类型", "攻击力", "描述"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            wb.save(self.items_file)
            print(f"📁 已创建物品文件: {self.items_file}")
    
    def _get_next_id(self):
        """获取下一个可用ID"""
        wb = openpyxl.load_workbook(self.items_file)
        ws = wb.active
        max_num = 0
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and str(cell_value).startswith("item_"):
                try:
                    num = int(str(cell_value).split("_")[1])
                    max_num = max(max_num, num)
                except:
                    continue
        return f"item_{str(max_num + 1).zfill(3)}"
    
    def add_item(self, name: str, item_type: str = "武器", attack: int = 0, description: str = ""):
        """新增物品"""
        try:
            wb = openpyxl.load_workbook(self.items_file)
            ws = wb.active
            
            item_id = self._get_next_id()
            next_row = ws.max_row + 1
            
            ws.cell(row=next_row, column=1, value=item_id)
            ws.cell(row=next_row, column=2, value=name)
            ws.cell(row=next_row, column=3, value=item_type)
            ws.cell(row=next_row, column=4, value=attack)
            ws.cell(row=next_row, column=5, value=description or f"一{item_type}")
            
            wb.save(self.items_file)
            
            return f"已添加物品 [{name}]，ID: {item_id}"
        except Exception as e:
            return f"添加失败: {str(e)}"
    
    def query_item(self, name_or_id: str):
        """查询物品"""
        try:
            wb = openpyxl.load_workbook(self.items_file)
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):
                item_id = ws.cell(row=row, column=1).value
                item_name = ws.cell(row=row, column=2).value
                
                if item_id == name_or_id or item_name == name_or_id:
                    data = []
                    for col in range(1, 6):
                        val = ws.cell(row=row, column=col).value
                        data.append(str(val) if val else "")
                    return f"物品信息: {' | '.join(data)}"
            
            return f"未找到物品: {name_or_id}"
        except Exception as e:
            return f"查询失败: {str(e)}"
    
    def list_items(self):
        """列出所有物品"""
        try:
            wb = openpyxl.load_workbook(self.items_file)
            ws = wb.active
            
            if ws.max_row <= 1:
                return "暂无物品"
            
            items = []
            for row in range(2, ws.max_row + 1):
                item_id = ws.cell(row=row, column=1).value
                name = ws.cell(row=row, column=2).value
                type_ = ws.cell(row=row, column=3).value
                attack = ws.cell(row=row, column=4).value
                if item_id:
                    items.append(f"{item_id} | {name} | {type_} | 攻击力:{attack}")
            
            return "📋 物品列表:\n" + "\n".join(items)
        except Exception as e:
            return f"列出失败: {str(e)}"