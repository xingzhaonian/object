from tkinter import Place
import easygui
import openpyxl
import os
import sys
import time

# 打开文件夹，获取表格绝对路径
def get_excel_file_path():
    g = easygui
    excel_file_path = g.fileopenbox(msg='请选择要转换的表格', title='选择表格')
    return excel_file_path

#打开文件夹 获取txt存放的路径
def get_txt_dir_path():
    g = easygui
    txt_dir_path = g.diropenbox(msg='选择要存放txt的目录', title='选择目录')
    return txt_dir_path

excel_file_path = get_excel_file_path()
txt_dir_path = get_txt_dir_path()
#print(excel_file_path, '===============', txt_dir_path)

# 进行转换操作
def excel_transformation_txt(path1, path2):
    time_num = 3
    try:
        txt_file_name = path1.split('\\')[-1].split('.')[0]
    except AttributeError:
        print('文件不存在')
        sys.exit()

    txt_file_path_name = txt_dir_path + '\\' + txt_file_name + '.txt'
    print(path1)
    print(path2)
    print(txt_file_path_name)
    wb = openpyxl.load_workbook(path1)
    sheet_name = wb.sheetnames
    print('表单名:', sheet_name)

    #获取表单名称
    while True:
        load_sheet_name  = input('请输入要转换的表单名称')
        try:
            sheet = wb[load_sheet_name]
        except KeyError:
            print('表单名输入错误, 请重新输入', load_sheet_name, sheet_name)
        finally:
            if load_sheet_name == sheet_name[0]:
                print('跳出循环')
                break
    sheet_max_row_num = sheet.max_row
    sheet_max_column_num = sheet.max_column
    print('表单有', sheet_max_row_num,'行', sheet_max_column_num, '列')
    sheet_row_data = sheet.rows
    row_data_num = 0


    #打开文件进行写入， 使用w+模式创建txt文本
    with open(txt_file_path_name, 'w+', encoding = 'utf-8') as f:
        for i in sheet_row_data:
            row_data_num += 1
            for k in i:
                k.value = str(k.value)      #转换单元格值的数据类型                                
                if k.value == 'None':
                    k.value = ''
                    print(row_data_num, type(k.value), k.value)
                f.write('\"' + k.value+'\"'+',')       #写入间隔符号
            f.write('\n')     #末尾进行换行
        f.close()
    # 打开多个文件进行符号和换行处理， 重新进行写入 
    m = txt_file_path_name.split('.')  #拆分路径
    k = m[0] + '1'                            
    result_path = k + '.' + m[1]
    with open(txt_file_path_name, 'r', encoding = 'utf-8')as a,\
     open(result_path, 'w', encoding = 'utf-8')as b: 
     while True:
        old_txt_data = a.readline()
        print(old_txt_data[0:-2])
        b.write(old_txt_data[0:-2])
        b.write('\n')
        if not old_txt_data:
            break
    #删除并重命名文件 
    os.remove(txt_file_path_name)
    os.rename(result_path, txt_file_path_name)


    for i in range(time_num):
        time_num -= 1
        time.sleep(time_num)
        print('写入完毕', time_num + 1, '秒后退出')

'''写入完毕'''
excel_transformation_txt(excel_file_path, txt_dir_path)
      
